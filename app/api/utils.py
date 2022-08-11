import json
import pandas as pd
import xlwings as xw
from string import ascii_uppercase
import xlsxwriter
import openpyxl 
from openpyxl import load_workbook,Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles.alignment import Alignment
import xlrd
import os
import math
import re
import io
import urllib3
from boto3 import client
from config import AWS_ACCESS_KEY_ID,AWS_SECRET_ACCESS_KEY,AWS_STORAGE_BUCKET_NAME,AWS_S3_REGION_NAME

def get_client():
    return client(
        's3',
        AWS_S3_REGION_NAME,
        aws_access_key_id=AWS_ACCESS_KEY_ID,
        aws_secret_access_key=AWS_SECRET_ACCESS_KEY
    )
    
def get_total_bytes(s3):
    result = s3.list_objects(Bucket=AWS_STORAGE_BUCKET_NAME)    
    for item in result['Contents']:
        if item['Key'] == 'org/1_1/81/1/sample_test1.xlsx':
            return item['Size']

## ***** accessories ***** ##
def get_cellrange(col_len):
    list_alphabets = [''] + [alp for alp in ascii_uppercase]
    list_columns = [str(j)+str(i) for j in list_alphabets for i in ascii_uppercase]
    return list_columns[:col_len]

def get_variable(varr):
    # return list of variables
    """
    matching patterns:  
    ${image_pos1,10,12,85%}, ${image_pos1$#2,10,12,85%}, ${bk_seq_num$#1}${style_fill_color$#1} ...
    ${style_fill_color}${style_fill_color}${style_fill_color}, ${style_fill_color}
    """ 
    variables = re.findall(r'\$\{\w+\$?\#?\d*,?\d*,?\d*,?\d*,?d*%?}', varr)  # match ${image_pos1,10,12,85%
    return variables

def get_variables_df(pxl):
    pxl.columns=get_cellrange(pxl.shape[1])
    df, variable_dict = pxl, {}
    rang, variable, variable1, c = [], [], [], 0
    for i in df.columns:
        indexss=list((df[(df[i].astype(str)).str.contains('\$')].reset_index(drop=False))['index'])
        if (len(indexss))>0:
            for indd in indexss:
                variable_dict[c]=str(i)+str(indd+1)+';'+str(df.loc[indd,str(i)])
                c+=1
    variable_location=pd.DataFrame.from_dict(variable_dict,orient='index').reset_index(drop=False).rename(columns={0:'range'})['range'].str.split(';',expand=True)
    variable_location.columns=['range','variable']

    for ind in variable_location.index:
        variables = get_variable(variable_location['variable'][ind])
        for var in variables:
            rang.append(variable_location['range'][ind])
            variable.append(variable_location['variable'][ind])
            variable1.append(var)
    variable_location = pd.DataFrame({"range": rang, 
                                "variable": variable,
                                "variable1": variable1})
    variable_location.drop_duplicates(inplace = True)
    return variable_location


def extract(templated_id,template_url):
    s3 = get_client()
    file = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME, Key=template_url)
    
    #doc_path = os.path.abspath(os.path.dirname(__file__)) + '\\template.xlsx'
    file_content = file['Body'].read()
    xl = load_workbook(io.BytesIO(file_content),data_only=True)    
    pxl=pd.read_excel(io.BytesIO(file_content), header=None)
    dict = {}
    sheet = xl.sheetnames[0]
    oxl = xl[sheet]
    pxl = pd.read_excel(io.BytesIO(file_content), sheet_name=sheet, header=None)
    variable_location = get_variables_df(pxl)
    list_var=list(variable_location['variable1'].unique())
    dict[str(sheet)] = list_var
    #variable_locations[sheet] = variable_location    
    
    x = {
        "template_id": templated_id,
        'change_key_list': dict
    }
    xl.close()
    return x

def replace_value(wb, ws, aoi, variable_location, image_loc, sheet_inc_value, image_info, clean):
    for i in aoi.keys():
        try:
            if clean:
                i_updated = i
            else:
                val = int(re.findall('#\d+', i)[0][1:]) - sheet_inc_value 
                i_updated = re.sub(r'#\d+', '#'+str(val), i)
        except:
            i_updated = i

        xxx=(variable_location[variable_location['variable1']==i_updated]).reset_index(drop=True)
        xxx_value=aoi[i]
        for k in xxx.index:
            cell_xl=xxx['range'][k]
            if len(xxx_value) < len(xxx) :
                xxx_value_str = str(xxx_value[0])
            else:
                xxx_value_str = str(xxx_value[k])
            if clean:
                ws[cell_xl].value = ""
            else:
                ws[cell_xl].value= ws[cell_xl].value.replace(str(xxx['variable1'][k]), xxx_value_str)
            #ws[cell_xl].value=str(xxx_value[k])
        if 'image' in i and xxx['range'].shape[0]>0:
            
            try:
                img_height = image_info[cell_xl][1]
                img_width = image_info[cell_xl][0]               
            except:
                img_height = 150
                img_width = 180
            try:
                png_loc="/content/pyxl1/"+str(xxx_value[0])
                file_exists = os.path.exists(png_loc)
            except:
                file_exists = 0
            if file_exists:
                my_png = openpyxl.drawing.image.Image(png_loc)
            else:
                try:
                    r = 1
                    http = urllib3.PoolManager()
                    r = http.request('GET', str(xxx_value[0]))
                    image_file = io.BytesIO(r.data)
                    my_png = openpyxl.drawing.image.Image(image_file)
                except:
                    my_png = "Not Found!"
            if my_png != 'Not Found!' and  not clean:
                my_png.height = img_height
                my_png.width = img_width
                ws[cell_xl].value = ""
                ws.add_image(my_png, str(xxx['range'][0]) )
                my_png.alignment = Alignment(horizontal='center',vertical='center')
                ws[str(xxx['range'][0])].alignment = Alignment(horizontal='center',vertical='center')

def process_step2(xl, replace_info, col_list, variable_location, image_info):
    replace_bukken_lists = replace_info['bukken_lists']
    total_sheets = math.ceil(len(replace_bukken_lists)/4)
    ws = xl[xl.sheetnames[0]]
    for _ in range(total_sheets-1):
        xl.copy_worksheet(ws)
    min_value = 0
    max_vale = 4
    sheet_inc_value = 0
    for sheet in xl.sheetnames:
        #replace_info = replace_info['replace_info']
        replace_header_info = {x: replace_info[x] for x in replace_info if x not in {'bukken_lists'}}
        ws = xl[sheet]
        images_loc = []
        for image in ws._images:
            row = image.anchor._from.row + 1
            col = col_list[image.anchor._from.col]
            images_loc.append(f'{col}{row}')
        if len(ws._images) > 0:
            xl_images_loc = list(variable_location[variable_location['variable1'].str.contains('image')]['range'])
            xl_images_indx = [images_loc.index(idx) for idx in xl_images_loc]
            ws._images = [ele for ele in ws._images if ele not in  list(map(ws._images.__getitem__, xl_images_indx))]
        replace_value(xl, ws, replace_header_info, variable_location, images_loc, sheet_inc_value, image_info, clean=False)
        replace_bukken_lists = replace_info['bukken_lists']
        area_of_intest = replace_bukken_lists[min_value:max_vale]
        for aoi in area_of_intest:
            replace_value(xl, ws, aoi, variable_location, images_loc, sheet_inc_value, image_info, clean = False)
        if len(area_of_intest)<4:
            if min_value>3:
                min_temp = min_value-4
                max_temp = max_vale-4
            for aoi in replace_bukken_lists[min_temp:max_temp][len(area_of_intest):]:
                replace_value(xl, ws, aoi, variable_location, images_loc, sheet_inc_value, image_info, clean = True)
        min_value += 4
        max_vale += 4
        sheet_inc_value += 4
    saved_file = 'template_output.xlsx'
    print(type(xl))
    xl.save(saved_file)
    print("Output File saved/updated successfuly!")
    print("Step 2 completed successfuly!")
    saved_url = saved_file
    return saved_file

def exchange(output_id,templated_id,template_url,replace_info):
    s3 = get_client()
    file = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME, Key=template_url)
    file_content = file['Body'].read()
    wb = load_workbook(io.BytesIO(file_content),data_only=True)    
    pxl=pd.read_excel(io.BytesIO(file_content), header=None)
    sheet = wb.sheetnames[0]
    
    pxl = pd.read_excel(io.BytesIO(file_content), sheet_name=sheet, header=None)
    variable_location = get_variables_df(pxl)
    
    # get all the info of the images in excel file
    image_info = {}
    ws = wb[wb.sheetnames[0]]
    pxl.columns=get_cellrange(pxl.shape[1])
    col_list = list(pxl.columns)
    merged_cells = ws.merged_cells.ranges
    rowHeights = [ws.row_dimensions[i+1].height for i in range(ws.max_row)]
    rowHeights = [15 if rh is None else rh for rh in rowHeights]
    for mc in merged_cells:
        col_min, row_min, col_max, row_max = mc.bounds
        col_width = 0
        for col in col_list[col_min-1:col_max]: col_width += ws.column_dimensions[col].width
        col_width *= 1.503
        row_height = (sum(rowHeights[row_min:row_max])) * 1.33
        image_info[mc.coord.split(":")[0]] = [col_width, row_height]
    
    wb = load_workbook(io.BytesIO(file_content),data_only=True)
    saved_file = process_step2(wb, replace_info, col_list, variable_location, image_info)
    wb.close()
    ### Upload excel file into S3
    out_template_url = str(template_url).split(".")[0]
    dest_filename = out_template_url + '_output.xlsx'
    print(dest_filename)
    
    s3.upload_file(saved_file,AWS_STORAGE_BUCKET_NAME,dest_filename)
    #url = create_presigned_url(AWS_STORAGE_BUCKET_NAME, dest_filename)
    response = {
        'output_id': output_id,
        'templated_id': templated_id,
        'report_url': dest_filename
    }
    return response